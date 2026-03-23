"""
Generate comprehensive pricing & profitability Excel workbook for Anvils.
Covers: money losers, money makers, per-customer profitability,
full service pricing matrix, optimal staffing model, and P&L projections.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ── Style definitions ──────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FONT = Font(name="Calibri", color="9C0006", bold=True)
GREEN_FONT = Font(name="Calibri", color="006100", bold=True)
BOLD = Font(name="Calibri", bold=True, size=11)
NORMAL = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
INR_FORMAT = '₹#,##0'
INR_FORMAT_DEC = '₹#,##0.00'
PCT_FORMAT = '0%'


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_row(ws, row, num_cols, fill=None, font=None):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font or NORMAL
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def auto_width(ws, min_width=12, max_width=45):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def add_section_header(ws, row, text, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", bold=True, size=13, color="2F5496")
    cell.alignment = Alignment(horizontal="left")
    return row + 1


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1: MONEY LOSERS
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Money Losers"

headers = ["Item", "Type", "Revenue to Anvils", "Cost to Anvils (Annual)", "Net P&L (Annual)", "Margin", "Strategic Purpose"]
row = 1
row = add_section_header(ws1, row, "MONEY LOSERS — Acquisition & Engagement Costs", len(headers))
row += 1

for c, h in enumerate(headers, 1):
    ws1.cell(row=row, column=c, value=h)
style_header_row(ws1, row, len(headers))
row += 1

losers = [
    ["Free Tier (per company)", "Platform", 0, 1200, -1200, -1.0, "Engagement engine. Drives marketplace purchases. 70% of companies stay here."],
    ["Incorporation — Launch Pvt Ltd", "Incorporation", 1499, 1000, 499, 0.33, "Wedge to get companies on platform. Not a profit center."],
    ["Incorporation — Launch OPC", "Incorporation", 999, 800, 199, 0.20, "Near break-even. Same wedge logic."],
    ["Incorporation — Launch LLP", "Incorporation", 1499, 900, 599, 0.40, "Competitive pricing for LLP acquisition."],
    ["GST Registration (service margin)", "Marketplace", 87, 0, 87, 1.0, "Rs 87 service margin is negligible. Purpose: onboard into GST monthly filing."],
    ["MSME / Udyam (service margin)", "Marketplace", 87, 0, 87, 1.0, "Free on govt portal. Service margin barely covers listing."],
    ["DIR-3 KYC (service margin)", "Marketplace", 87, 0, 87, 1.0, "Low value per director. Volume play."],
    ["Starter Subscription", "SaaS", 4999, 4800, 199, 0.04, "Break-even at best. Value gap between Free and Starter is small."],
]

for item in losers:
    for c, val in enumerate(item, 1):
        ws1.cell(row=row, column=c, value=val)
    # Color code based on net P&L
    net = item[4]
    fill = RED_FILL if net < 0 else YELLOW_FILL if net < 500 else None
    font = RED_FONT if net < 0 else None
    style_data_row(ws1, row, len(headers), fill=fill, font=font)
    # Format currency columns
    for col in [3, 4, 5]:
        ws1.cell(row=row, column=col).number_format = INR_FORMAT
    ws1.cell(row=row, column=6).number_format = PCT_FORMAT
    row += 1

# Total row
row += 1
ws1.cell(row=row, column=1, value="TOTAL ACQUISITION COST PER COMPANY (Year 1)")
ws1.cell(row=row, column=1).font = BOLD
ws1.cell(row=row, column=5, value=-1200 + 499)  # Free tier loss + incorporation margin
ws1.cell(row=row, column=5).number_format = INR_FORMAT
ws1.cell(row=row, column=5).font = RED_FONT
ws1.cell(row=row, column=7, value="CAC = Rs 700-3,000 per company (free tier cost + thin incorp margin)")
ws1.cell(row=row, column=7).font = BOLD

auto_width(ws1)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2: MONEY MAKERS
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Money Makers")

# --- Tier 1: Subscriptions ---
headers2 = ["Item", "Type", "Annual Revenue", "Annual Cost to Serve", "Gross Profit", "Margin", "Notes"]
row = 1
row = add_section_header(ws2, row, "TIER 1: SUBSCRIPTIONS (85-92% Margin — The Real Business)", len(headers2))
row += 1
for c, h in enumerate(headers2, 1):
    ws2.cell(row=row, column=c, value=h)
style_header_row(ws2, row, len(headers2))
row += 1

subs = [
    ["Growth Subscription", "SaaS", 29999, 5000, 24999, 0.83, "Target: funded startups. 40-70% cheaper than Trica/Qapita."],
    ["Scale Subscription", "SaaS", 99999, 10000, 89999, 0.90, "Target: Series A+. 60-85% cheaper than Carta."],
]
for item in subs:
    for c, val in enumerate(item, 1):
        ws2.cell(row=row, column=c, value=val)
    style_data_row(ws2, row, len(headers2), fill=GREEN_FILL, font=GREEN_FONT)
    for col in [3, 4, 5]:
        ws2.cell(row=row, column=col).number_format = INR_FORMAT
    ws2.cell(row=row, column=6).number_format = PCT_FORMAT
    row += 1

# Scale target
row += 1
ws2.cell(row=row, column=1, value="Month 24 Target: 200 Growth + 30 Scale")
ws2.cell(row=row, column=1).font = BOLD
ws2.cell(row=row, column=3, value=200*29999 + 30*99999)
ws2.cell(row=row, column=3).number_format = INR_FORMAT
ws2.cell(row=row, column=5, value=200*24999 + 30*89999)
ws2.cell(row=row, column=5).number_format = INR_FORMAT
ws2.cell(row=row, column=7, value="Rs 89.9L/yr subscription revenue, Rs 77.7L/yr gross profit")
row += 2

# --- Tier 2: High-Frequency Marketplace ---
row = add_section_header(ws2, row, "TIER 2: HIGH-FREQUENCY MARKETPLACE (Service Margin × Volume × Repeat)", len(headers2))
row += 1
headers_freq = ["Service", "Type", "Customer Pays/yr", "Anvils Service Margin/yr (17.5%)", "Frequency", "Notes", ""]
for c, h in enumerate(headers_freq, 1):
    ws2.cell(row=row, column=c, value=h)
style_header_row(ws2, row, len(headers_freq))
row += 1

freq_services = [
    ["Bookkeeping Standard (12mo)", "Accounting", 71988, 12598, "Monthly", "Highest service margin per customer"],
    ["Bookkeeping Basic (12mo)", "Accounting", 35988, 6298, "Monthly", "High-volume SME service"],
    ["Payroll Processing (12mo)", "Accounting", 23988, 4198, "Monthly", "Scales with employee count"],
    ["TDS Quarterly (4x)", "Tax", 9996, 1749, "Quarterly", "Mandatory for most companies"],
    ["GST Monthly Filing (12mo)", "Tax", 9588, 1678, "Monthly", "Highest-frequency service"],
]
for item in freq_services:
    for c, val in enumerate(item, 1):
        ws2.cell(row=row, column=c, value=val)
    style_data_row(ws2, row, len(headers_freq), fill=GREEN_FILL)
    for col in [3, 4]:
        ws2.cell(row=row, column=col).number_format = INR_FORMAT
    row += 1

row += 1
ws2.cell(row=row, column=1, value="Typical Pvt Ltd (GST + TDS + Basic Bookkeeping)")
ws2.cell(row=row, column=1).font = BOLD
ws2.cell(row=row, column=4, value=1678 + 1749 + 6298)
ws2.cell(row=row, column=4).number_format = INR_FORMAT
ws2.cell(row=row, column=6, value="Rs 9,725/yr service margin per customer at ZERO fulfillment cost")
row += 2

# --- Tier 3: High-Ticket Marketplace ---
row = add_section_header(ws2, row, "TIER 3: HIGH-TICKET MARKETPLACE (Fewer Events, Higher Service Margin)", len(headers2))
row += 1
headers_ticket = ["Service", "Type", "Customer Pays", "Anvils Service Margin (17.5%)", "When", "Notes", ""]
for c, h in enumerate(headers_ticket, 1):
    ws2.cell(row=row, column=c, value=h)
style_header_row(ws2, row, len(headers_ticket))
row += 1

ticket_services = [
    ["Statutory Audit", "Tax", 14999, 2625, "Annual (mandatory)", "Biggest single-event service margin"],
    ["Company Closure", "Amendment", 9999, 1750, "One-time", "Complex, multi-month process"],
    ["Share Allotment", "Amendment", 8999, 1575, "During funding", "Platform-adjacent (cap table)"],
    ["Increase Capital", "Amendment", 8999, 1575, "During funding", "Often bundled with allotment"],
    ["Annual ROC Filing", "Compliance", 7999, 1400, "Annual (mandatory)", "Every Pvt Ltd/OPC must file"],
    ["Share Transfer", "Amendment", 7999, 1400, "Equity events", "Platform-adjacent (cap table)"],
    ["Name Change", "Amendment", 7999, 1400, "One-time", "Rare but premium priced"],
]
for item in ticket_services:
    for c, val in enumerate(item, 1):
        ws2.cell(row=row, column=c, value=val)
    style_data_row(ws2, row, len(headers_ticket), fill=GREEN_FILL)
    for col in [3, 4]:
        ws2.cell(row=row, column=col).number_format = INR_FORMAT
    row += 1

# --- Tier 4: Incorporation (Grow/Scale tiers) ---
row += 2
row = add_section_header(ws2, row, "TIER 4: INCORPORATION — Grow/Scale Tiers (Actual Margin)", len(headers2))
row += 1
headers_inc = ["Entity + Tier", "Type", "Platform Fee", "Est. Cost", "Profit", "Margin", ""]
for c, h in enumerate(headers_inc, 1):
    ws2.cell(row=row, column=c, value=h)
style_header_row(ws2, row, len(headers_inc))
row += 1

inc_tiers = [
    ["Pvt Ltd — Grow", "Incorporation", 3999, 1000, 2999, 0.75],
    ["Pvt Ltd — Scale", "Incorporation", 7999, 1500, 6499, 0.81],
    ["LLP — Grow", "Incorporation", 3499, 1000, 2499, 0.71],
    ["LLP — Scale", "Incorporation", 6999, 1500, 5499, 0.79],
    ["OPC — Grow", "Incorporation", 2999, 800, 2199, 0.73],
    ["OPC — Scale", "Incorporation", 5999, 1200, 4799, 0.80],
]
for item in inc_tiers:
    for c, val in enumerate(item, 1):
        ws2.cell(row=row, column=c, value=val)
    style_data_row(ws2, row, len(headers_inc), fill=GREEN_FILL)
    for col in [3, 4, 5]:
        ws2.cell(row=row, column=col).number_format = INR_FORMAT
    ws2.cell(row=row, column=6).number_format = PCT_FORMAT
    row += 1

auto_width(ws2)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3: PER-CUSTOMER PROFITABILITY
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Per-Customer Profit")

headers3 = ["Revenue Line", "Bootstrapped (Free Tier)", "Funded Startup (Growth)", "Series A+ (Scale)"]
row = 1
row = add_section_header(ws3, row, "PER-CUSTOMER PROFITABILITY — Year 1 (Pvt Ltd)", len(headers3))
row += 1
for c, h in enumerate(headers3, 1):
    ws3.cell(row=row, column=c, value=h)
style_header_row(ws3, row, len(headers3))
row += 1

cust_data = [
    ["Incorporation (platform fee)", 3999, 7999, 7999],
    ["Subscription (annual)", 0, 29999, 99999],
    ["Marketplace service margin — mandatory filings", 9200, 9200, 9200],
    ["Marketplace service margin — equity events", 0, 4550, 4550],
    ["TOTAL ANVILS REVENUE", 13199, 51748, 121748],
    ["", "", "", ""],
    ["Cost to serve (infra + support)", 1800, 6000, 12000],
    ["GROSS PROFIT", 11399, 45748, 109748],
    ["GROSS MARGIN", 0.86, 0.88, 0.90],
]

for item in cust_data:
    for c, val in enumerate(item, 1):
        ws3.cell(row=row, column=c, value=val)
    if item[0] in ["TOTAL ANVILS REVENUE", "GROSS PROFIT"]:
        style_data_row(ws3, row, len(headers3), fill=GREEN_FILL, font=BOLD)
    elif item[0] == "GROSS MARGIN":
        for col in [2, 3, 4]:
            ws3.cell(row=row, column=col).number_format = PCT_FORMAT
        style_data_row(ws3, row, len(headers3), font=BOLD)
    else:
        style_data_row(ws3, row, len(headers3))
    for col in [2, 3, 4]:
        if item[0] != "GROSS MARGIN" and item[0] != "":
            ws3.cell(row=row, column=col).number_format = INR_FORMAT
    row += 1

# LTV section
row += 2
row = add_section_header(ws3, row, "CUSTOMER LIFETIME VALUE (3-Year)", len(headers3))
row += 1
for c, h in enumerate(headers3, 1):
    ws3.cell(row=row, column=c, value=h)
style_header_row(ws3, row, len(headers3))
row += 1

ltv_data = [
    ["Year 1 Revenue", 13199, 51748, 121748],
    ["Year 2 Revenue", 9200, 43149, 113149],
    ["Year 3 Revenue", 9200, 43149, 113149],
    ["3-Year LTV", 31599, 138046, 348046],
    ["Retention Rate (Annual)", 0.65, 0.82, 0.92],
    ["Adjusted 3-Year LTV", 23799, 119178, 327163],
]

for item in ltv_data:
    for c, val in enumerate(item, 1):
        ws3.cell(row=row, column=c, value=val)
    if "LTV" in str(item[0]):
        style_data_row(ws3, row, len(headers3), fill=GREEN_FILL, font=BOLD)
    elif "Retention" in str(item[0]):
        for col in [2, 3, 4]:
            ws3.cell(row=row, column=col).number_format = PCT_FORMAT
        style_data_row(ws3, row, len(headers3), font=BOLD)
    else:
        style_data_row(ws3, row, len(headers3))
    for col in [2, 3, 4]:
        if "Retention" not in str(item[0]):
            ws3.cell(row=row, column=col).number_format = INR_FORMAT
    row += 1

auto_width(ws3)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4: FULL SERVICE PRICING MATRIX
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Service Pricing Matrix")

headers4 = ["#", "Service", "Category", "Current Fee", "Market Low", "Market High",
            "Proposed Fee", "Action", "Strategy Bucket", "Anvils Platform Margin (17.5%)",
            "Position vs Market", "Rationale"]
row = 1
row = add_section_header(ws4, row, "FULL SERVICE PRICING MATRIX — All 37 Services", len(headers4))
row += 1
for c, h in enumerate(headers4, 1):
    ws4.cell(row=row, column=c, value=h)
style_header_row(ws4, row, len(headers4))
row += 1

services = [
    # Registration
    [1, "GST Registration", "Registration", 499, 499, 1999, 499, "HOLD", "Wedge", 87, "At floor", "Entry point to GST filing revenue"],
    [2, "MSME / Udyam", "Registration", 499, 499, 2899, 499, "HOLD", "Wedge", 87, "Below market", "Free on govt portal"],
    [3, "Trademark Registration", "Registration", 4999, 1899, 4999, 3499, "REDUCE", "Anchor", 612, "Midpoint", "Was at ceiling; IndiaFilings at 1,899"],
    [4, "IEC Code", "Registration", 1999, 1499, 3999, 1999, "HOLD", "Anchor", 350, "Low-mid", "Niche, low volume"],
    [5, "FSSAI Basic", "Registration", 2499, 1899, 3499, 2499, "HOLD", "Anchor", 437, "Midpoint", "Food businesses"],
    [6, "FSSAI State License", "Registration", 5999, 4999, 9999, 5999, "HOLD", "Premium", 1050, "Low end", "Complex service"],
    [7, "DPIIT Startup India", "Registration", 2999, 1999, 4999, 2999, "HOLD", "Anchor", 525, "Midpoint", "Free on govt portal"],
    [8, "Professional Tax", "Registration", 1499, 999, 2499, 1499, "HOLD", "Anchor", 262, "Midpoint", "State-level"],
    [9, "ESI Registration", "Registration", 2499, 1499, 3999, 2499, "HOLD", "Anchor", 437, "Midpoint", "10+ employees"],
    [10, "EPFO / PF Registration", "Registration", 2499, 1499, 3999, 2499, "HOLD", "Anchor", 437, "Midpoint", "20+ employees"],
    [11, "ISO 9001 Certification", "Registration", 19999, 15000, 40000, 19999, "HOLD", "Premium", 3500, "Low end", "Enterprise service"],
    # Compliance
    [12, "Annual ROC Filing", "Compliance", 7999, 5000, 15000, 7999, "HOLD", "Anchor", 1400, "Midpoint", "Mandatory for all companies"],
    [13, "LLP Annual Filing", "Compliance", 5999, 3299, 8000, 4999, "REDUCE", "Anchor", 875, "Midpoint", "Was above mid; LegalWiz at 3,299"],
    [14, "DIR-3 KYC", "Compliance", 499, 300, 1000, 499, "HOLD", "Anchor", 87, "Midpoint", "Per-director volume play"],
    [15, "ADT-1 Auditor Appt", "Compliance", 1999, 1500, 3000, 1999, "HOLD", "Anchor", 350, "Midpoint", "Bundled with audit"],
    [16, "INC-20A Commencement", "Compliance", 1999, 1500, 3500, 1999, "HOLD", "Wedge", 350, "Low-mid", "Year 1 only"],
    # Tax
    [17, "ITR-6 (Company)", "Tax", 4999, 3000, 15000, 4999, "HOLD", "Anchor", 875, "Lower mid", "CAs charge 5K-15K"],
    [18, "ITR-5 (LLP)", "Tax", 2999, 2000, 8000, 2999, "HOLD", "Anchor", 525, "Lower mid", "Competitive"],
    [19, "ITR Individual", "Tax", 999, 500, 3000, 999, "HOLD", "Wedge", 175, "Lower mid", "Low absolute value"],
    [20, "GST Monthly Filing", "Tax", 999, 400, 2500, 799, "REDUCE", "Frequency", 140, "Competitive", "Highest-freq; MyOnlineCA at 400"],
    [21, "GST Annual Return", "Tax", 4999, 3000, 10000, 4999, "HOLD", "Anchor", 875, "Midpoint", "Annual filing"],
    [22, "TDS Quarterly", "Tax", 2499, 1500, 5000, 2499, "HOLD", "Frequency", 437, "Midpoint", "4x/yr; competitive vs LegalWiz"],
    [23, "Statutory Audit", "Tax", 14999, 10000, 50000, 14999, "HOLD", "Premium", 2625, "Low end", "Quote-based; starting price"],
    # Accounting
    [24, "Bookkeeping Basic", "Accounting", 2999, 1500, 5000, 2999, "HOLD", "Frequency", 525, "Midpoint", "Monthly, SLA-backed"],
    [25, "Bookkeeping Standard", "Accounting", 5999, 4000, 10000, 5999, "HOLD", "Premium", 1050, "Lower mid", "Higher-volume businesses"],
    [26, "Payroll Processing", "Accounting", 1999, 1000, 5000, 1999, "HOLD", "Frequency", 350, "Lower mid", "Scales with headcount"],
    # Amendments
    [27, "Director Change", "Amendment", 3499, 2999, 7999, 4999, "RAISE", "Premium", 875, "40th pctl", "Event-driven, 30-day deadline"],
    [28, "Share Transfer", "Amendment", 4999, 3999, 9999, 7999, "RAISE", "Premium", 1400, "Midpoint", "Platform-adjacent: cap table"],
    [29, "Share Allotment", "Amendment", 5999, 4999, 12000, 8999, "RAISE", "Premium", 1575, "Above mid", "During fundraising rounds"],
    [30, "Increase Capital", "Amendment", 5999, 5499, 15000, 8999, "RAISE", "Premium", 1575, "Below mid", "Precedes share allotment"],
    [31, "Office Change", "Amendment", 3499, 2999, 7999, 4999, "RAISE", "Premium", 875, "40th pctl", "Event-driven"],
    [32, "Name Change", "Amendment", 5999, 4999, 9999, 7999, "RAISE", "Premium", 1400, "60th pctl", "Rare, premium execution"],
    [33, "Company Closure", "Amendment", 7999, 7999, 15000, 9999, "RAISE", "Premium", 1750, "Below mid", "Most complex service"],
    [34, "LLP Partner Change", "Amendment", 3499, 2999, 5999, 3999, "RAISE", "Premium", 700, "Midpoint", "LLPs price-sensitive"],
    # Legal
    [35, "Trademark Objection", "Legal", 4999, 3999, 9999, 4999, "HOLD", "Premium", 875, "Low end", "IP attorney required"],
    [36, "Legal Notice", "Legal", 3499, 2999, 7999, 3499, "HOLD", "Premium", 612, "Low end", "Advocate-dependent"],
    [37, "Virtual Office", "Legal", 7999, 7999, 14000, 7999, "HOLD", "Wedge", 1400, "At floor", "43% cheaper than IndiaFilings"],
]

for item in services:
    for c, val in enumerate(item, 1):
        ws4.cell(row=row, column=c, value=val)
    action = item[7]
    if action == "REDUCE":
        fill = YELLOW_FILL
    elif action == "RAISE":
        fill = GREEN_FILL
    else:
        fill = None
    style_data_row(ws4, row, len(headers4), fill=fill)
    for col in [4, 5, 6, 7, 10]:
        ws4.cell(row=row, column=col).number_format = INR_FORMAT
    row += 1

# Summary
row += 1
ws4.cell(row=row, column=1, value="SUMMARY")
ws4.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=13, color="2F5496")
row += 1
summary_items = [
    ["Price Reductions", 3, "Trademark, LLP Annual, GST Monthly"],
    ["Price Increases", 8, "All amendment/event-based services"],
    ["Hold", 26, "Well-positioned services"],
    ["Net Revenue Impact", "", "Positive — increases on premium services exceed reductions on frequency services"],
]
for item in summary_items:
    for c, val in enumerate(item, 1):
        ws4.cell(row=row, column=c, value=val)
    ws4.cell(row=row, column=1).font = BOLD
    row += 1

auto_width(ws4)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5: STAFFING MODEL
# ═══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Staffing Model")

row = 1
row = add_section_header(ws5, row, "OPTIMAL STAFFING TO MAXIMIZE PROFIT", 8)
row += 1

# Key insight box
ws5.cell(row=row, column=1, value="KEY INSIGHT: Under the marketplace model, Anvils does NOT employ CAs/CSs for ongoing compliance work.")
ws5.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11, color="9C0006")
row += 1
ws5.cell(row=row, column=1, value="Marketplace CAs handle all filing fulfillment. Anvils only needs staff for: incorporation filings, platform ops, customer success, and marketing.")
ws5.cell(row=row, column=1).font = NORMAL
row += 2

# Phase 1: Lean Launch
headers5 = ["Role", "Headcount", "Monthly CTC", "Annual CTC", "Revenue They Enable (Annual)", "ROI", "Notes"]
row = add_section_header(ws5, row, "PHASE 1: LEAN LAUNCH (Month 1-6) — 3 People, Rs 2.75L/mo burn", len(headers5))
row += 1
for c, h in enumerate(headers5, 1):
    ws5.cell(row=row, column=c, value=h)
style_header_row(ws5, row, len(headers5))
row += 1

phase1 = [
    ["Founder (CEO + Sales + Ops)", 1, 50000, 600000, "All revenue", "∞", "Product direction, CA partnerships, sales"],
    ["Full-Stack Developer", 1, 100000, 1200000, "Platform = all revenue", "∞", "Build and maintain platform"],
    ["Contract Customer Success", 1, 50000, 600000, 900000, "1.5x", "Onboarding, retention, marketplace coordination"],
    ["", "", "", "", "", "", ""],
    ["TOTAL PHASE 1", 3, 200000, 2400000, "", "", "Break-even possible at Month 12-14"],
]
for item in phase1:
    for c, val in enumerate(item, 1):
        ws5.cell(row=row, column=c, value=val)
    if item[0] == "TOTAL PHASE 1":
        style_data_row(ws5, row, len(headers5), fill=SUBHEADER_FILL, font=BOLD)
    else:
        style_data_row(ws5, row, len(headers5))
    for col in [3, 4]:
        if isinstance(ws5.cell(row=row, column=col).value, (int, float)):
            ws5.cell(row=row, column=col).number_format = INR_FORMAT
    row += 1

row += 1

# Phase 2: Growth
row = add_section_header(ws5, row, "PHASE 2: GROWTH (Month 7-12) — 5 People, Rs 3.75L/mo burn", len(headers5))
row += 1
for c, h in enumerate(headers5, 1):
    ws5.cell(row=row, column=c, value=h)
style_header_row(ws5, row, len(headers5))
row += 1

phase2 = [
    ["Founder", 1, 50000, 600000, "", "", "Same role"],
    ["Full-Stack Developer", 1, 100000, 1200000, "", "", "Same role"],
    ["Contract Customer Success", 1, 50000, 600000, "", "", "Same role"],
    ["Ops Manager (Incorp + Marketplace)", 1, 60000, 720000, 1500000, "2.1x", "Incorporation processing, CA partner QA"],
    ["Content Writer / Marketing", 1, 50000, 600000, 800000, "1.3x", "SEO, blog, social media, PLG tools"],
    ["", "", "", "", "", "", ""],
    ["TOTAL PHASE 2", 5, 310000, 3720000, "", "", "Revenue should be Rs 1.25-3.35L/mo"],
]
for item in phase2:
    for c, val in enumerate(item, 1):
        ws5.cell(row=row, column=c, value=val)
    if item[0] == "TOTAL PHASE 2":
        style_data_row(ws5, row, len(headers5), fill=SUBHEADER_FILL, font=BOLD)
    else:
        style_data_row(ws5, row, len(headers5))
    for col in [3, 4, 5]:
        if isinstance(ws5.cell(row=row, column=col).value, (int, float)):
            ws5.cell(row=row, column=col).number_format = INR_FORMAT
    row += 1

row += 1

# Phase 3
row = add_section_header(ws5, row, "PHASE 3: SCALE (Month 13-24) — 8-9 People, Rs 5-6L/mo burn", len(headers5))
row += 1
for c, h in enumerate(headers5, 1):
    ws5.cell(row=row, column=c, value=h)
style_header_row(ws5, row, len(headers5))
row += 1

phase3 = [
    ["Founder", 1, 50000, 600000, "", "", "Strategy + enterprise sales"],
    ["Full-Stack Developer (Senior)", 1, 100000, 1200000, "", "", "Core platform"],
    ["Full-Stack Developer (Mid)", 1, 80000, 960000, "", "", "Mobile app, integrations"],
    ["Customer Success Manager", 1, 60000, 720000, "", "", "Growth/Scale tier onboarding + upsell"],
    ["Customer Success Associate", 1, 30000, 360000, "", "", "Support, follow-ups"],
    ["Ops Manager", 1, 60000, 720000, "", "", "Incorporation + marketplace QA"],
    ["Content Writer / Marketing", 1, 50000, 600000, "", "", "SEO + content"],
    ["Business Development", 1, 50000, 600000, 2000000, "3.3x", "Enterprise sales for Scale tier"],
    ["Finance (part-time)", 0.5, 25000, 300000, "", "", "Internal bookkeeping, payroll"],
    ["", "", "", "", "", "", ""],
    ["TOTAL PHASE 3", 8.5, 505000, 6060000, "", "", "Revenue target: Rs 5-16L/mo"],
]
for item in phase3:
    for c, val in enumerate(item, 1):
        ws5.cell(row=row, column=c, value=val)
    if item[0] == "TOTAL PHASE 3":
        style_data_row(ws5, row, len(headers5), fill=SUBHEADER_FILL, font=BOLD)
    else:
        style_data_row(ws5, row, len(headers5))
    for col in [3, 4, 5]:
        if isinstance(ws5.cell(row=row, column=col).value, (int, float)):
            ws5.cell(row=row, column=col).number_format = INR_FORMAT
    row += 1

# Comparison section
row += 2
row = add_section_header(ws5, row, "STAFFING COMPARISON: LEAN vs FULL OPS MODEL", 8)
row += 1
comp_headers = ["Metric", "Lean Model (Recommended)", "Full Ops Model (Old)", "Savings", "", "", "", ""]
for c, h in enumerate(comp_headers, 1):
    ws5.cell(row=row, column=c, value=h)
style_header_row(ws5, row, 8)
row += 1

comparisons = [
    ["Launch headcount", 3, 11.5, "8.5 fewer people"],
    ["Launch monthly burn", 275000, 1020000, "Rs 7.45L/mo saved"],
    ["Year 1 salary cost", 2400000, 9200000, "Rs 68L saved"],
    ["Year 1 total cost", 3300000, 15000000, "Rs 1.17 Cr saved"],
    ["Break-even month", "Month 12-14", "Month 14-16", "2-4 months faster"],
    ["Capital required", 3000000, 15000000, "Rs 1.2 Cr less capital needed"],
    ["Biggest difference", "No in-house CA/CS team", "6 CA/CS on payroll", "Marketplace CAs do the work"],
]
for item in comparisons:
    for c, val in enumerate(item, 1):
        ws5.cell(row=row, column=c, value=val)
    style_data_row(ws5, row, 8)
    for col in [2, 3]:
        if isinstance(ws5.cell(row=row, column=col).value, (int, float)):
            ws5.cell(row=row, column=col).number_format = INR_FORMAT
    row += 1

# Hiring rules
row += 2
row = add_section_header(ws5, row, "HIRING RULES — Never Hire Ahead of Revenue", 8)
row += 1
rules = [
    ["RULE 1", "Do not hire until monthly revenue exceeds 50% of new monthly burn after the hire"],
    ["RULE 2", "Never hire a CA or CS as an employee — use marketplace partners instead"],
    ["RULE 3", "Gate Phase 2 hires on: 30+ incorporations/mo AND 65+ Starter subscribers"],
    ["RULE 4", "Gate Phase 3 hires on: 60+ incorporations/mo AND 15+ Growth subscribers"],
    ["RULE 5", "Outsource bookkeeping always — thin margins, scales linearly, no strategic value"],
    ["RULE 6", "First non-founder hire should be Ops Manager (incorporation volume is the bottleneck)"],
    ["RULE 7", "Second non-founder hire should be Content Writer (SEO is highest-ROI acquisition channel)"],
    ["RULE 8", "Hire Customer Success Manager only when Growth subscribers exceed 10 (upsell justifies cost)"],
]
for item in rules:
    ws5.cell(row=row, column=1, value=item[0])
    ws5.cell(row=row, column=1).font = BOLD
    ws5.cell(row=row, column=2, value=item[1])
    ws5.cell(row=row, column=2).font = NORMAL
    row += 1

auto_width(ws5)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6: P&L PROJECTION
# ═══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("P&L Projection")

row = 1
row = add_section_header(ws6, row, "18-MONTH P&L PROJECTION (Lean Model)", 14)
row += 1

pl_headers = ["", "M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11", "M12"]
for c, h in enumerate(pl_headers, 1):
    ws6.cell(row=row, column=c, value=h)
style_header_row(ws6, row, len(pl_headers))
row += 1

# Revenue rows
revenue_data = [
    ["REVENUE", "", "", "", "", "", "", "", "", "", "", "", ""],
    ["Incorporations", 10, 10, 15, 15, 20, 30, 30, 35, 40, 45, 50, 60],
    ["Incorp Revenue", 25000, 25000, 37500, 37500, 50000, 75000, 75000, 87500, 100000, 112500, 125000, 150000],
    ["Starter Subs (cumulative)", 8, 15, 25, 35, 48, 65, 80, 95, 112, 130, 150, 175],
    ["Growth Subs", 0, 0, 0, 1, 2, 4, 6, 8, 10, 12, 15, 17],
    ["Scale Subs", 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1, 2],
    ["SaaS Revenue", 3992, 7485, 12475, 20465, 29950, 44431, 57920, 71413, 95878, 110862, 129835, 158308],
    ["Marketplace Revenue", 0, 600, 1200, 2400, 3600, 6000, 7800, 11050, 13000, 18200, 21000, 27300],
    ["TOTAL REVENUE", 28992, 33085, 51175, 60365, 83550, 125431, 140720, 169963, 208878, 241562, 275835, 335608],
    ["", "", "", "", "", "", "", "", "", "", "", "", ""],
    ["COSTS", "", "", "", "", "", "", "", "", "", "", "", ""],
    ["Salaries (team)", 200000, 200000, 200000, 200000, 200000, 200000, 310000, 310000, 310000, 310000, 310000, 310000],
    ["Technology", 15000, 15000, 15000, 15000, 15000, 15000, 18000, 18000, 18000, 18000, 18000, 18000],
    ["Marketing", 25000, 25000, 25000, 25000, 25000, 25000, 30000, 30000, 30000, 30000, 30000, 30000],
    ["Office + Misc", 15000, 15000, 15000, 15000, 15000, 15000, 17000, 17000, 17000, 17000, 17000, 17000],
    ["Contingency (10%)", 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000, 20000],
    ["TOTAL COSTS", 275000, 275000, 275000, 275000, 275000, 275000, 395000, 395000, 395000, 395000, 395000, 395000],
    ["", "", "", "", "", "", "", "", "", "", "", "", ""],
    ["NET P&L", -246008, -241915, -223825, -214635, -191450, -149569, -254280, -225037, -186122, -153438, -119165, -59392],
    ["CUMULATIVE P&L", -246008, -487923, -711748, -926383, -1117833, -1267402, -1521682, -1746719, -1932841, -2086279, -2205444, -2264836],
]

for item in revenue_data:
    for c, val in enumerate(item, 1):
        ws6.cell(row=row, column=c, value=val)
    label = item[0]
    if label in ["REVENUE", "COSTS"]:
        style_data_row(ws6, row, len(pl_headers), fill=SUBHEADER_FILL, font=BOLD)
    elif label in ["TOTAL REVENUE"]:
        style_data_row(ws6, row, len(pl_headers), fill=GREEN_FILL, font=BOLD)
    elif label in ["TOTAL COSTS"]:
        style_data_row(ws6, row, len(pl_headers), fill=RED_FILL, font=BOLD)
    elif label == "NET P&L":
        style_data_row(ws6, row, len(pl_headers), font=RED_FONT)
        for col in range(2, len(pl_headers) + 1):
            ws6.cell(row=row, column=col).number_format = INR_FORMAT
    elif label == "CUMULATIVE P&L":
        style_data_row(ws6, row, len(pl_headers), font=RED_FONT)
        for col in range(2, len(pl_headers) + 1):
            ws6.cell(row=row, column=col).number_format = INR_FORMAT
    else:
        style_data_row(ws6, row, len(pl_headers))

    # Format currency rows
    if label in ["Incorp Revenue", "SaaS Revenue", "Marketplace Revenue", "TOTAL REVENUE",
                 "Salaries (team)", "Technology", "Marketing", "Office + Misc",
                 "Contingency (10%)", "TOTAL COSTS"]:
        for col in range(2, len(pl_headers) + 1):
            ws6.cell(row=row, column=col).number_format = INR_FORMAT
    row += 1

# Key metrics
row += 1
ws6.cell(row=row, column=1, value="Year 1 Total Revenue")
ws6.cell(row=row, column=1).font = BOLD
ws6.cell(row=row, column=2, value=sum(revenue_data[8][1:]))  # TOTAL REVENUE row
ws6.cell(row=row, column=2).number_format = INR_FORMAT
row += 1
ws6.cell(row=row, column=1, value="Year 1 Total Cost")
ws6.cell(row=row, column=1).font = BOLD
ws6.cell(row=row, column=2, value=sum(revenue_data[16][1:]))  # TOTAL COSTS row
ws6.cell(row=row, column=2).number_format = INR_FORMAT
row += 1
ws6.cell(row=row, column=1, value="Year 1 Net Loss")
ws6.cell(row=row, column=1).font = BOLD
ws6.cell(row=row, column=2, value=revenue_data[19][-1])  # Last cumulative
ws6.cell(row=row, column=2).number_format = INR_FORMAT
ws6.cell(row=row, column=2).font = RED_FONT
row += 1
ws6.cell(row=row, column=1, value="M12 Revenue as % of Burn")
ws6.cell(row=row, column=1).font = BOLD
ws6.cell(row=row, column=2, value=round(335608 / 395000, 2))
ws6.cell(row=row, column=2).number_format = PCT_FORMAT
row += 1
ws6.cell(row=row, column=1, value="Estimated Break-Even")
ws6.cell(row=row, column=1).font = BOLD
ws6.cell(row=row, column=2, value="Month 13-14")

auto_width(ws6)

# ═══════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════
output_path = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..",
    "business",
    "anvils-pricing-profitability.xlsx"
)
output_path = os.path.normpath(output_path)
wb.save(output_path)
print(f"Saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
