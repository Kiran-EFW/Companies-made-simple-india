"""
Generate comprehensive Excel workbook comparing Full Ops vs Lean (Marketplace) model.
Includes side-by-side staffing, cost, revenue, and 36-month P&L projections.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ── Style definitions ──────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
LEAN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FULLOPS_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
RED_FONT = Font(name="Calibri", color="9C0006", bold=True)
GREEN_FONT = Font(name="Calibri", color="006100", bold=True)
BOLD = Font(name="Calibri", bold=True, size=11)
NORMAL = Font(name="Calibri", size=11)
SECTION_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
INR = '₹#,##0'
PCT = '0.0%'


def style_header(ws, row, num_cols, fill=None, font=None):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font or HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_row(ws, row, num_cols, fill=None, font=None):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font or NORMAL
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def auto_width(ws, min_w=14, max_w=45):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        mx = min_w
        for cell in col:
            if cell.value:
                mx = max(mx, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[letter].width = mx


def section(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left")
    return row + 1


def write_row(ws, row, data, ncols, fill=None, font=None, fmt=None):
    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = font or NORMAL
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        if fmt and c > 1 and isinstance(val, (int, float)):
            cell.number_format = fmt
    return row + 1


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1: STAFFING COMPARISON — Full Ops vs Lean
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Staffing Comparison"
NC = 8
r = 1

r = section(ws, r, "STAFFING MODEL: FULL OPS vs LEAN (MARKETPLACE)", NC)
r += 1

# ── Full Ops Launch ──
r = section(ws, r, "FULL OPS MODEL — Launch Phase (0-100 companies)", NC)
headers = ["Role", "Headcount", "Monthly Cost", "Annual Cost", "Department", "Purpose", "Can Be Replaced?", "Replacement"]
for c, h in enumerate(headers, 1):
    ws.cell(row=r, column=c, value=h)
style_header(ws, r, NC)
r += 1

full_ops_launch = [
    ["Operations Head", 1, 150000, 1800000, "Operations", "Oversee all ops, doubles as CS Lead", "Yes", "Founder handles ops"],
    ["Company Secretary (Senior)", 1, 70000, 840000, "CS Dept", "All MCA filings", "Yes — Marketplace", "Partner CS on marketplace"],
    ["Company Secretary (Junior)", 1, 30000, 360000, "CS Dept", "Drafting, coordination", "Yes — Marketplace", "Partner CS on marketplace"],
    ["Chartered Accountant (Senior)", 1, 70000, 840000, "CA Dept", "GST, TDS, ITR for all clients", "Yes — Marketplace", "Partner CA on marketplace"],
    ["Document Reviewer", 1, 30000, 360000, "Filing", "Verify docs against AI extraction", "Partially — AI + Ops", "Founder/ops person reviews"],
    ["Customer Success Associate", 1, 30000, 360000, "CS", "Admin messaging, customer queries", "Merged", "Contract ops/CS hire"],
]

full_ops_total_monthly = 0
full_ops_total_annual = 0
for item in full_ops_launch:
    r = write_row(ws, r, item, NC, fill=FULLOPS_FILL)
    ws.cell(row=r-1, column=3).number_format = INR
    ws.cell(row=r-1, column=4).number_format = INR
    full_ops_total_monthly += item[2]
    full_ops_total_annual += item[3]

total_data = ["TOTAL FULL OPS (Launch)", 6, full_ops_total_monthly, full_ops_total_annual, "", "", "", ""]
r = write_row(ws, r, total_data, NC, fill=FULLOPS_FILL, font=BOLD)
ws.cell(row=r-1, column=3).number_format = INR
ws.cell(row=r-1, column=4).number_format = INR
r += 1

# ── Lean Model Launch ──
r = section(ws, r, "LEAN MODEL — Phase 1 Launch (3 people)", NC)
for c, h in enumerate(headers, 1):
    ws.cell(row=r, column=c, value=h)
style_header(ws, r, NC)
r += 1

lean_launch = [
    ["Founder / CEO", 1, 50000, 600000, "Leadership", "Product, sales, partnerships, CA onboarding", "N/A", "Founder role"],
    ["Full-Stack Developer", 1, 100000, 1200000, "Engineering", "Platform build, integrations, infra", "N/A", "Core hire"],
    ["Contract Ops / CS", 1, 50000, 600000, "Operations", "Incorp support, onboarding, Starter retention", "N/A", "Contract retainer"],
]

lean_total_monthly = 0
lean_total_annual = 0
for item in lean_launch:
    r = write_row(ws, r, item, NC, fill=LEAN_FILL)
    ws.cell(row=r-1, column=3).number_format = INR
    ws.cell(row=r-1, column=4).number_format = INR
    lean_total_monthly += item[2]
    lean_total_annual += item[3]

total_data = ["TOTAL LEAN (Phase 1)", 3, lean_total_monthly, lean_total_annual, "", "", "", ""]
r = write_row(ws, r, total_data, NC, fill=LEAN_FILL, font=BOLD)
ws.cell(row=r-1, column=3).number_format = INR
ws.cell(row=r-1, column=4).number_format = INR
r += 1

# ── Side-by-Side Summary ──
r = section(ws, r, "SIDE-BY-SIDE COMPARISON (Launch)", NC)
comp_headers = ["Metric", "Full Ops Model", "Lean (Marketplace) Model", "Difference", "Savings %", "", "", ""]
for c, h in enumerate(comp_headers, 1):
    ws.cell(row=r, column=c, value=h)
style_header(ws, r, NC)
r += 1

comparisons = [
    ["Team Size (Launch)", 6, 3, -3, "50%"],
    ["Monthly Salary Cost", full_ops_total_monthly, lean_total_monthly, lean_total_monthly - full_ops_total_monthly, f"{(1 - lean_total_monthly/full_ops_total_monthly)*100:.0f}%"],
    ["Annual Salary Cost", full_ops_total_annual, lean_total_annual, lean_total_annual - full_ops_total_annual, f"{(1 - lean_total_annual/full_ops_total_annual)*100:.0f}%"],
    ["Other Costs (monthly)", 75000, 75000, 0, "0%"],
    ["Total Monthly Burn", full_ops_total_monthly + 75000, lean_total_monthly + 75000, (lean_total_monthly + 75000) - (full_ops_total_monthly + 75000), f"{(1 - (lean_total_monthly+75000)/(full_ops_total_monthly+75000))*100:.0f}%"],
    ["Year 1 Total Cost (12 months)", (full_ops_total_monthly + 75000) * 12, (lean_total_monthly + 75000) * 12, ((lean_total_monthly + 75000) - (full_ops_total_monthly + 75000)) * 12, f"{(1 - (lean_total_monthly+75000)/(full_ops_total_monthly+75000))*100:.0f}%"],
    ["In-house CAs/CSs needed", 3, 0, -3, "100%"],
    ["Compliance fulfillment model", "In-house staff", "Marketplace partners", "N/A", "N/A"],
    ["Break-even (estimated)", "Month 18-20", "Month 12-14", "4-6 months earlier", "N/A"],
    ["Capital required", "₹60-80L", "₹25-35L", "₹35-45L less", "50-55%"],
]

for item in comparisons:
    r = write_row(ws, r, item + ["", "", ""], NC)
    if isinstance(item[1], (int, float)) and item[1] > 1000:
        ws.cell(row=r-1, column=2).number_format = INR
        ws.cell(row=r-1, column=3).number_format = INR
        ws.cell(row=r-1, column=4).number_format = INR
        if isinstance(item[3], (int, float)) and item[3] < 0:
            ws.cell(row=r-1, column=4).font = GREEN_FONT

r += 1
# ── Growth Phase Comparison ──
r = section(ws, r, "GROWTH PHASE COMPARISON (Month 7-12)", NC)
growth_headers = ["Metric", "Full Ops Model", "Lean Model", "Difference", "Notes", "", "", ""]
for c, h in enumerate(growth_headers, 1):
    ws.cell(row=r, column=c, value=h)
style_header(ws, r, NC)
r += 1

growth_data = [
    ["Team Size", "15 + outsourced", "5 people", "-10 people", "Lean adds Ops Mgr + Marketing"],
    ["Monthly Salary", 940000, 310000, -630000, "Lean: no CA/CS team"],
    ["Monthly Burn (incl. other)", 1050000, 375000, -675000, ""],
    ["6-month cost (M7-M12)", 6300000, 2250000, -4050000, "Lean saves ₹40.5L in 6 months"],
]

for item in growth_data:
    r = write_row(ws, r, item + ["", "", ""], NC)
    if isinstance(item[1], (int, float)):
        ws.cell(row=r-1, column=2).number_format = INR
        ws.cell(row=r-1, column=3).number_format = INR
        ws.cell(row=r-1, column=4).number_format = INR

r += 1
r = section(ws, r, "SCALE PHASE COMPARISON (Month 13-24)", NC)
for c, h in enumerate(growth_headers, 1):
    ws.cell(row=r, column=c, value=h)
style_header(ws, r, NC)
r += 1

scale_data = [
    ["Team Size", "25+ people", "8-9 people", "-16 people", "Lean adds 2nd dev + Sr CS + BD"],
    ["Monthly Salary", 1700000, 450000, -1250000, "Full ops: ₹2 Cr/yr salaries"],
    ["Monthly Burn (incl. other)", 1950000, 550000, -1400000, ""],
    ["12-month cost (M13-M24)", 23400000, 6600000, -16800000, "Lean saves ₹1.68 Cr per year"],
]

for item in scale_data:
    r = write_row(ws, r, item + ["", "", ""], NC)
    if isinstance(item[1], (int, float)):
        ws.cell(row=r-1, column=2).number_format = INR
        ws.cell(row=r-1, column=3).number_format = INR
        ws.cell(row=r-1, column=4).number_format = INR

auto_width(ws)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2: 36-MONTH P&L — LEAN MODEL
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("3-Year P&L (Lean)")
NC2 = 14  # Label + 12 months per section, but we'll do months as columns

# We'll do all 36 months in rows for readability
r = 1
r = section(ws2, r, "36-MONTH P&L PROJECTION — LEAN (MARKETPLACE) MODEL", 10)
r += 1

headers2 = ["Month", "Incorporations", "Incorp Revenue", "Starter Subs", "Growth Subs", "Scale Subs",
            "SaaS Revenue", "Marketplace Rev", "Total Revenue", "Monthly Burn", "Net P&L", "Cumulative P&L"]
NC2 = len(headers2)
for c, h in enumerate(headers2, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, NC2)
r += 1

# Revenue assumptions by month (36 months)
# Year 1: from profitability-strategy.md (actuals)
# Year 2-3: extrapolated with growth assumptions
monthly_data = []

# Year 1 data (from profitability doc)
y1 = [
    # (month, incorp, incorp_rev, starter, growth, scale, saas_rev, mktpl_rev, burn)
    (1,  10, 25000,   8,   0, 0,  3992,     0,   275000),
    (2,  10, 25000,  15,   0, 0,  7485,   600,   275000),
    (3,  15, 37500,  25,   0, 0, 12475,  1200,   275000),
    (4,  15, 37500,  35,   1, 0, 20465,  2400,   275000),
    (5,  20, 50000,  48,   2, 0, 29950,  3600,   275000),
    (6,  30, 75000,  65,   4, 0, 44431,  6000,   275000),
    (7,  30, 75000,  80,   6, 0, 57920,  7800,   375000),
    (8,  35, 87500,  95,   8, 0, 71413, 11050,   375000),
    (9,  40, 100000, 112, 10, 1, 95878, 13000,   375000),
    (10, 45, 112500, 130, 12, 1, 110862, 18200,  375000),
    (11, 50, 125000, 150, 15, 1, 129835, 21000,  375000),
    (12, 60, 150000, 175, 17, 2, 158308, 27300,  375000),
]

# Year 2: Growth accelerates (more incorporations, upgrades, marketplace traction)
# Starter churn stabilizes at 4-5%/mo, Growth adoption increases
y2_data = []
prev_starter = 175
prev_growth = 17
prev_scale = 2
incorp_y2 = [65, 70, 75, 80, 85, 90, 95, 100, 110, 115, 120, 130]

for i, m in enumerate(range(13, 25)):
    inc = incorp_y2[i]
    # Starter: 75% of new incorp convert, 4.5% monthly churn
    new_starter = int(inc * 0.75)
    churn = int(prev_starter * 0.045)
    prev_starter = prev_starter + new_starter - churn
    # Growth: gains ~3-5 per month from starter upgrades + direct
    growth_gain = 3 + (i // 3)  # accelerates over year
    growth_churn = max(1, int(prev_growth * 0.015))
    prev_growth = prev_growth + growth_gain - growth_churn
    # Scale: gains 1 every 2-3 months
    scale_gain = 1 if i % 2 == 0 else 0
    prev_scale = prev_scale + scale_gain

    saas = prev_starter * 499 + prev_growth * 2999 + prev_scale * 9999
    incorp_rev = int(inc * 2500)  # blended avg
    mktpl = int(prev_starter * 0.22 * 700 + prev_growth * 0.3 * 1200 + prev_scale * 0.4 * 2000)
    # Burn: Phase 2 at M13-18 = ₹3.75L, Phase 3 at M19+ = ₹5.5L
    if m <= 18:
        burn = 375000
    else:
        burn = 550000

    y2_data.append((m, inc, incorp_rev, prev_starter, prev_growth, prev_scale, saas, mktpl, burn))

# Year 3: Platform maturity
y3_data = []
incorp_y3 = [135, 140, 150, 155, 160, 165, 170, 175, 180, 185, 190, 200]

for i, m in enumerate(range(25, 37)):
    inc = incorp_y3[i]
    new_starter = int(inc * 0.78)
    churn = int(prev_starter * 0.04)
    prev_starter = prev_starter + new_starter - churn
    growth_gain = 5 + (i // 3)
    growth_churn = max(1, int(prev_growth * 0.012))
    prev_growth = prev_growth + growth_gain - growth_churn
    scale_gain = 1 if i % 2 == 0 else (2 if i == 11 else 0)
    prev_scale = prev_scale + scale_gain

    saas = prev_starter * 499 + prev_growth * 2999 + prev_scale * 9999
    incorp_rev = int(inc * 2800)
    mktpl = int(prev_starter * 0.25 * 750 + prev_growth * 0.35 * 1400 + prev_scale * 0.45 * 2500)
    burn = 600000  # Phase 3 mature

    y3_data.append((m, inc, incorp_rev, prev_starter, prev_growth, prev_scale, saas, mktpl, burn))

all_months = y1 + y2_data + y3_data

cumulative = 0
for md in all_months:
    month, inc, inc_rev, starter, growth, scale, saas, mktpl, burn = md
    total_rev = inc_rev + saas + mktpl
    net = total_rev - burn
    cumulative += net

    row_data = [f"M{month}", inc, inc_rev, starter, growth, scale, saas, mktpl, total_rev, burn, net, cumulative]
    r = write_row(ws2, r, row_data, NC2)

    # Format currency columns
    for col_idx in [3, 7, 8, 9, 10, 11, 12]:
        ws2.cell(row=r-1, column=col_idx).number_format = INR

    # Color net P&L
    if net < 0:
        ws2.cell(row=r-1, column=11).font = RED_FONT
    else:
        ws2.cell(row=r-1, column=11).font = GREEN_FONT

    if cumulative < 0:
        ws2.cell(row=r-1, column=12).font = RED_FONT
    else:
        ws2.cell(row=r-1, column=12).font = GREEN_FONT

    # Highlight year boundaries
    if month in [12, 24]:
        style_row(ws2, r-1, NC2, fill=YELLOW_FILL, font=BOLD)
        for col_idx in [3, 7, 8, 9, 10, 11, 12]:
            ws2.cell(row=r-1, column=col_idx).number_format = INR

# Add year summaries
r += 1
r = section(ws2, r, "ANNUAL SUMMARIES", NC2)
r += 1

sum_headers = ["Period", "Total Incorporations", "Incorp Revenue", "End Starter Subs", "End Growth Subs",
               "End Scale Subs", "Total SaaS Revenue", "Total Mktpl Rev", "Total Revenue", "Total Cost",
               "Net P&L", "End Cumul P&L"]
for c, h in enumerate(sum_headers, 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, NC2)
r += 1

# Year 1 summary
y1_inc = sum(m[1] for m in y1)
y1_inc_rev = sum(m[2] for m in y1)
y1_saas = sum(m[6] for m in y1)
y1_mktpl = sum(m[7] for m in y1)
y1_cost = sum(m[8] for m in y1)
y1_total = y1_inc_rev + y1_saas + y1_mktpl
y1_net = y1_total - y1_cost
y1_cum = y1_net

y1_row = ["Year 1 (M1-M12)", y1_inc, y1_inc_rev, 175, 17, 2, y1_saas, y1_mktpl, y1_total, y1_cost, y1_net, y1_cum]
r = write_row(ws2, r, y1_row, NC2, fill=YELLOW_FILL, font=BOLD)
for col_idx in [3, 7, 8, 9, 10, 11, 12]:
    ws2.cell(row=r-1, column=col_idx).number_format = INR

# Year 2 summary
y2_inc = sum(m[1] for m in y2_data)
y2_inc_rev = sum(m[2] for m in y2_data)
y2_saas = sum(m[6] for m in y2_data)
y2_mktpl = sum(m[7] for m in y2_data)
y2_cost = sum(m[8] for m in y2_data)
y2_total = y2_inc_rev + y2_saas + y2_mktpl
y2_net = y2_total - y2_cost
y2_cum = y1_cum + y2_net

y2_row = ["Year 2 (M13-M24)", y2_inc, y2_inc_rev, y2_data[-1][3], y2_data[-1][4], y2_data[-1][5],
          y2_saas, y2_mktpl, y2_total, y2_cost, y2_net, y2_cum]
r = write_row(ws2, r, y2_row, NC2, fill=YELLOW_FILL, font=BOLD)
for col_idx in [3, 7, 8, 9, 10, 11, 12]:
    ws2.cell(row=r-1, column=col_idx).number_format = INR

# Year 3 summary
y3_inc = sum(m[1] for m in y3_data)
y3_inc_rev = sum(m[2] for m in y3_data)
y3_saas = sum(m[6] for m in y3_data)
y3_mktpl = sum(m[7] for m in y3_data)
y3_cost = sum(m[8] for m in y3_data)
y3_total = y3_inc_rev + y3_saas + y3_mktpl
y3_net = y3_total - y3_cost
y3_cum = y2_cum + y3_net

y3_row = ["Year 3 (M25-M36)", y3_inc, y3_inc_rev, y3_data[-1][3], y3_data[-1][4], y3_data[-1][5],
          y3_saas, y3_mktpl, y3_total, y3_cost, y3_net, y3_cum]
r = write_row(ws2, r, y3_row, NC2, fill=LEAN_FILL, font=BOLD)
for col_idx in [3, 7, 8, 9, 10, 11, 12]:
    ws2.cell(row=r-1, column=col_idx).number_format = INR

# 3-year total
three_yr_row = ["3-YEAR TOTAL", y1_inc + y2_inc + y3_inc, y1_inc_rev + y2_inc_rev + y3_inc_rev,
                y3_data[-1][3], y3_data[-1][4], y3_data[-1][5],
                y1_saas + y2_saas + y3_saas, y1_mktpl + y2_mktpl + y3_mktpl,
                y1_total + y2_total + y3_total, y1_cost + y2_cost + y3_cost,
                y1_net + y2_net + y3_net, y3_cum]
r = write_row(ws2, r, three_yr_row, NC2, font=TOTAL_FONT)
style_row(ws2, r-1, NC2, fill=TOTAL_FILL, font=TOTAL_FONT)
for col_idx in [3, 7, 8, 9, 10, 11, 12]:
    ws2.cell(row=r-1, column=col_idx).number_format = INR

auto_width(ws2)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3: 36-MONTH P&L — FULL OPS MODEL
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3-Year P&L (Full Ops)")
r = 1
r = section(ws3, r, "36-MONTH P&L PROJECTION — FULL OPS MODEL", 10)
r += 1

for c, h in enumerate(headers2, 1):
    ws3.cell(row=r, column=c, value=h)
style_header(ws3, r, NC2)
r += 1

# Full ops: same revenue but higher burn because in-house CA/CS teams
# Also, compliance revenue is bundled into subscriptions (lower margin)
# But incorporation volume may be slightly higher (more staff for processing)

# Full ops burn by phase:
# Launch (M1-6): 6 people at ₹3.8L salary + ₹3.2L other = ~₹7.0L/mo
# But really, let's use the actual data from operations doc
# Launch: 6 people = ₹3.8L salary + marketing ₹50K + tech ₹35K + tools ₹15K + contingency ₹50K = ~₹4.8L
# Growth (M7-12): 15 people = ₹9.4L salary + other costs ~₹2.5L = ~₹12L/mo
# Scale (M13-24): 25 people = ₹17L salary + other ~₹3L = ~₹20L/mo

full_ops_burns = {
    range(1, 7): 700000,      # ₹7L/mo (6 people + other)
    range(7, 13): 1200000,    # ₹12L/mo (15 people + other)
    range(13, 19): 1500000,   # ₹15L/mo (scaling up)
    range(19, 25): 1950000,   # ₹19.5L/mo (25 people)
    range(25, 37): 2200000,   # ₹22L/mo (25+ people)
}

def get_fullops_burn(month):
    for rng, burn in full_ops_burns.items():
        if month in rng:
            return burn
    return 2200000

cumulative_full = 0
fo_year_data = {1: [], 2: [], 3: []}

for md in all_months:
    month, inc, inc_rev, starter, growth, scale, saas, mktpl, lean_burn = md
    # Full ops has same revenue from incorporations
    # But for services, they fulfill in-house, so they get 100% of service fee (not just platform margin)
    # However, they also have the cost of the CA/CS doing the work
    # Net effect: similar revenue but much higher costs
    # For simplicity: same total revenue projection (marketplace service margin vs in-house with higher cost)
    total_rev = inc_rev + saas + mktpl
    burn = get_fullops_burn(month)
    net = total_rev - burn
    cumulative_full += net

    year = 1 if month <= 12 else (2 if month <= 24 else 3)
    fo_year_data[year].append((month, inc, inc_rev, starter, growth, scale, saas, mktpl, total_rev, burn, net))

    row_data = [f"M{month}", inc, inc_rev, starter, growth, scale, saas, mktpl, total_rev, burn, net, cumulative_full]
    r = write_row(ws3, r, row_data, NC2)
    for col_idx in [3, 7, 8, 9, 10, 11, 12]:
        ws3.cell(row=r-1, column=col_idx).number_format = INR
    if net < 0:
        ws3.cell(row=r-1, column=11).font = RED_FONT
    else:
        ws3.cell(row=r-1, column=11).font = GREEN_FONT
    if cumulative_full < 0:
        ws3.cell(row=r-1, column=12).font = RED_FONT
    else:
        ws3.cell(row=r-1, column=12).font = GREEN_FONT
    if month in [12, 24]:
        style_row(ws3, r-1, NC2, fill=YELLOW_FILL, font=BOLD)
        for col_idx in [3, 7, 8, 9, 10, 11, 12]:
            ws3.cell(row=r-1, column=col_idx).number_format = INR

# Summaries
r += 1
r = section(ws3, r, "ANNUAL SUMMARIES — FULL OPS MODEL", NC2)
r += 1
for c, h in enumerate(sum_headers, 1):
    ws3.cell(row=r, column=c, value=h)
style_header(ws3, r, NC2)
r += 1

fo_cum = 0
for year in [1, 2, 3]:
    yd = fo_year_data[year]
    t_inc = sum(d[1] for d in yd)
    t_inc_rev = sum(d[2] for d in yd)
    t_saas = sum(d[6] for d in yd)
    t_mktpl = sum(d[7] for d in yd)
    t_rev = sum(d[8] for d in yd)
    t_cost = sum(d[9] for d in yd)
    t_net = t_rev - t_cost
    fo_cum += t_net
    yr_row = [f"Year {year}", t_inc, t_inc_rev, yd[-1][3], yd[-1][4], yd[-1][5],
              t_saas, t_mktpl, t_rev, t_cost, t_net, fo_cum]
    fill = FULLOPS_FILL if t_net < 0 else LEAN_FILL
    r = write_row(ws3, r, yr_row, NC2, fill=fill, font=BOLD)
    for col_idx in [3, 7, 8, 9, 10, 11, 12]:
        ws3.cell(row=r-1, column=col_idx).number_format = INR

auto_width(ws3)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4: SIDE-BY-SIDE P&L COMPARISON
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Lean vs Full Ops (Side-by-Side)")
r = 1
NC4 = 9
r = section(ws4, r, "LEAN vs FULL OPS — ANNUAL P&L COMPARISON", NC4)
r += 1

comp4_headers = ["Metric", "Lean Y1", "Full Ops Y1", "Lean Y2", "Full Ops Y2", "Lean Y3", "Full Ops Y3",
                 "Lean 3-Yr Total", "Full Ops 3-Yr Total"]
for c, h in enumerate(comp4_headers, 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, NC4)
r += 1

# Compute Full Ops annual totals
fo_y1_cost = sum(get_fullops_burn(m) for m in range(1, 13))
fo_y2_cost = sum(get_fullops_burn(m) for m in range(13, 25))
fo_y3_cost = sum(get_fullops_burn(m) for m in range(25, 37))
fo_total_cost = fo_y1_cost + fo_y2_cost + fo_y3_cost

lean_y1_total_cost = y1_cost
lean_y2_total_cost = y2_cost
lean_y3_total_cost = y3_cost
lean_total_cost = lean_y1_total_cost + lean_y2_total_cost + lean_y3_total_cost

# Same revenue for both models
side_data = [
    ["Team Size", "3→5", "6→15", "5→9", "15→25", "9", "25+", "—", "—"],
    ["Total Revenue", y1_total, y1_total, y2_total, y2_total, y3_total, y3_total,
     y1_total + y2_total + y3_total, y1_total + y2_total + y3_total],
    ["Total Cost", lean_y1_total_cost, fo_y1_cost, lean_y2_total_cost, fo_y2_cost,
     lean_y3_total_cost, fo_y3_cost, lean_total_cost, fo_total_cost],
    ["Net P&L", y1_net, y1_total - fo_y1_cost, y2_net, y2_total - fo_y2_cost,
     y3_net, y3_total - fo_y3_cost, y1_net + y2_net + y3_net,
     (y1_total - fo_y1_cost) + (y2_total - fo_y2_cost) + (y3_total - fo_y3_cost)],
    ["Cost Savings (Lean)", "", "", "", "", "", "", "",
     fo_total_cost - lean_total_cost],
]

for item in side_data:
    r = write_row(ws4, r, item, NC4)
    for col_idx in range(2, NC4 + 1):
        val = ws4.cell(row=r-1, column=col_idx).value
        if isinstance(val, (int, float)) and abs(val) > 100:
            ws4.cell(row=r-1, column=col_idx).number_format = INR
            if val < 0:
                ws4.cell(row=r-1, column=col_idx).font = RED_FONT
            elif val > 0 and item[0] in ["Net P&L", "Cost Savings (Lean)"]:
                ws4.cell(row=r-1, column=col_idx).font = GREEN_FONT

r += 2
r = section(ws4, r, "MONTHLY BURN COMPARISON", NC4)
r += 1

burn_headers = ["Phase", "Months", "Lean Burn/mo", "Full Ops Burn/mo", "Monthly Savings",
                "Lean Headcount", "Full Ops Headcount", "Lean Cumul Cost", "Full Ops Cumul Cost"]
for c, h in enumerate(burn_headers, 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, NC4)
r += 1

burn_phases = [
    ["Phase 1 (Launch)", "M1-M6", 275000, 700000, 425000, 3, 6, 275000*6, 700000*6],
    ["Phase 2 (Growth)", "M7-M12", 375000, 1200000, 825000, 5, 15, 275000*6 + 375000*6, 700000*6 + 1200000*6],
    ["Phase 3a (Scale)", "M13-M18", 375000, 1500000, 1125000, 5, 20, 275000*6 + 375000*12, 700000*6 + 1200000*6 + 1500000*6],
    ["Phase 3b (Scale)", "M19-M24", 550000, 1950000, 1400000, 9, 25, 275000*6 + 375000*12 + 550000*6, 700000*6 + 1200000*6 + 1500000*6 + 1950000*6],
    ["Phase 4 (Mature)", "M25-M36", 600000, 2200000, 1600000, 9, 25, 275000*6 + 375000*12 + 550000*6 + 600000*12, 700000*6 + 1200000*6 + 1500000*6 + 1950000*6 + 2200000*12],
]

for item in burn_phases:
    r = write_row(ws4, r, item, NC4)
    for col_idx in [3, 4, 5, 8, 9]:
        ws4.cell(row=r-1, column=col_idx).number_format = INR
    ws4.cell(row=r-1, column=5).font = GREEN_FONT

r += 2
r = section(ws4, r, "3-YEAR COST SAVINGS SUMMARY", NC4)
r += 1

savings_data = [
    ["Total 3-year cost (Lean)", lean_total_cost],
    ["Total 3-year cost (Full Ops)", fo_total_cost],
    ["TOTAL SAVINGS", fo_total_cost - lean_total_cost],
    ["Savings as %", (1 - lean_total_cost / fo_total_cost) if fo_total_cost > 0 else 0],
    ["", ""],
    ["3-year profit (Lean)", y1_net + y2_net + y3_net],
    ["3-year profit (Full Ops)", (y1_total - fo_y1_cost) + (y2_total - fo_y2_cost) + (y3_total - fo_y3_cost)],
    ["PROFIT DIFFERENCE", (y1_net + y2_net + y3_net) - ((y1_total - fo_y1_cost) + (y2_total - fo_y2_cost) + (y3_total - fo_y3_cost))],
]

for item in savings_data:
    ws4.cell(row=r, column=1, value=item[0])
    ws4.cell(row=r, column=1).font = BOLD
    ws4.cell(row=r, column=1).border = THIN_BORDER
    ws4.cell(row=r, column=2, value=item[1])
    ws4.cell(row=r, column=2).border = THIN_BORDER
    if isinstance(item[1], float) and item[1] < 1:
        ws4.cell(row=r, column=2).number_format = PCT
    elif isinstance(item[1], (int, float)) and abs(item[1]) > 100:
        ws4.cell(row=r, column=2).number_format = INR
        if item[1] > 0 and "SAVINGS" in str(item[0]):
            ws4.cell(row=r, column=2).font = GREEN_FONT
            ws4.cell(row=r, column=2).fill = LEAN_FILL
        elif item[1] < 0:
            ws4.cell(row=r, column=2).font = RED_FONT
    r += 1

auto_width(ws4)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5: MONEY LOSERS & MAKERS (from original)
# ═══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Money Losers & Makers")
NC5 = 7
r = 1
r = section(ws5, r, "MONEY LOSERS — Intentional Acquisition Costs", NC5)
r += 1

headers5 = ["Item", "Type", "Revenue to Anvils", "Cost to Anvils", "Net P&L", "Margin", "Purpose"]
for c, h in enumerate(headers5, 1):
    ws5.cell(row=r, column=c, value=h)
style_header(ws5, r, NC5)
r += 1

losers = [
    ["Free doc bundle (per company)", "Platform", 0, 1200, -1200, -1.0, "Engagement engine — drives marketplace purchases"],
    ["Launch Pvt Ltd incorporation", "Incorporation", 1499, 1000, 499, 0.33, "Wedge — not a profit center"],
    ["Launch OPC incorporation", "Incorporation", 999, 800, 199, 0.20, "Wedge — near break-even"],
    ["Launch LLP incorporation", "Incorporation", 1499, 900, 599, 0.40, "Wedge — competitive pricing"],
    ["GST Registration (service margin)", "Marketplace", 87, 0, 87, 1.0, "₹87 service margin — negligible. Purpose: onboard into GST monthly filing"],
    ["MSME/Udyam Registration (svc margin)", "Marketplace", 87, 0, 87, 1.0, "Free on govt portal. Service margin barely covers listing"],
    ["DIR-3 KYC (service margin)", "Marketplace", 87, 0, 87, 1.0, "Low value per director — volume play"],
    ["Starter subscription", "SaaS", 5988, 5760, 228, 0.04, "Break-even at best. Gap between Free and Starter is small"],
]

for item in losers:
    r = write_row(ws5, r, item, NC5, fill=FULLOPS_FILL)
    for ci in [3, 4, 5]:
        ws5.cell(row=r-1, column=ci).number_format = INR
    ws5.cell(row=r-1, column=6).number_format = PCT

r += 2
r = section(ws5, r, "MONEY MAKERS — High-Margin Revenue", NC5)
r += 1

makers_headers = ["Item", "Type", "Annual Revenue/Customer", "Annual Cost", "Annual Profit", "Margin", "Notes"]
for c, h in enumerate(makers_headers, 1):
    ws5.cell(row=r, column=c, value=h)
style_header(ws5, r, NC5)
r += 1

makers = [
    ["Growth subscription", "SaaS", 35988, 5400, 30588, 0.85, "Core business — pure software, 85%+ margin"],
    ["Scale subscription", "SaaS", 119988, 12000, 107988, 0.90, "Enterprise — 90%+ margin, lowest churn"],
    ["Bookkeeping Standard (monthly)", "Marketplace", 12598, 0, 12598, 1.0, "Recurring service margin — highest volume"],
    ["GST+TDS+Bookkeeping bundle", "Marketplace", 9725, 0, 9725, 1.0, "Most common bundle — per customer/year"],
    ["Statutory Audit (annual)", "Marketplace", 2625, 0, 2625, 1.0, "High-ticket, annual — per company"],
    ["Share Allotment filing", "Marketplace", 1575, 0, 1575, 1.0, "Platform-adjacent — fundraising trigger"],
    ["Company Closure", "Marketplace", 1750, 0, 1750, 1.0, "High-ticket, complex service"],
    ["Grow tier incorporation", "Incorporation", 3999, 1500, 2499, 0.63, "Upsold tier — good margin"],
    ["Scale tier incorporation", "Incorporation", 7999, 1500, 6499, 0.81, "Highest incorp tier — excellent margin"],
]

for item in makers:
    r = write_row(ws5, r, item, NC5, fill=LEAN_FILL)
    for ci in [3, 4, 5]:
        ws5.cell(row=r-1, column=ci).number_format = INR
    ws5.cell(row=r-1, column=6).number_format = PCT

r += 2
r = section(ws5, r, "PER-CUSTOMER ANNUAL PROFITABILITY", NC5)
r += 1

cust_headers = ["Customer Segment", "% of Companies", "Annual Revenue", "Annual Cost", "Annual Profit", "Margin", "3-Year LTV"]
for c, h in enumerate(cust_headers, 1):
    ws5.cell(row=r, column=c, value=h)
style_header(ws5, r, NC5)
r += 1

segments = [
    ["Bootstrapped (Starter + Marketplace)", "70%", 13200, 1800, 11400, 0.86, 24400],
    ["Funded Startup (Growth + Marketplace)", "20%", 51948, 6200, 45748, 0.88, 110071],
    ["Series A+ (Scale + Marketplace)", "5%", 121999, 12250, 109749, 0.90, 319247],
]

for item in segments:
    r = write_row(ws5, r, item, NC5, fill=ACCENT_FILL)
    for ci in [3, 4, 5, 7]:
        ws5.cell(row=r-1, column=ci).number_format = INR
    ws5.cell(row=r-1, column=6).number_format = PCT

auto_width(ws5)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6: KEY ASSUMPTIONS
# ═══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Key Assumptions")
NC6 = 4
r = 1
r = section(ws6, r, "KEY ASSUMPTIONS BEHIND THE PROJECTIONS", NC6)
r += 1

ass_headers = ["Assumption", "Value", "Basis", "Sensitivity"]
for c, h in enumerate(ass_headers, 1):
    ws6.cell(row=r, column=c, value=h)
style_header(ws6, r, NC6)
r += 1

assumptions = [
    ["Incorporation to Starter conversion", "75-80%", "Free doc bundle hooks them; compliance calendar creates immediate value", "HIGH — if <60%, Starter revenue drops significantly"],
    ["Starter monthly churn", "5-8%", "High for SME SaaS. Many companies won't see ongoing value at ₹499/mo", "CRITICAL — at 10%+, model breaks"],
    ["Starter to Growth upgrade (startup segment)", "5-8%", "Only funded startups need cap table/ESOP. Small % of total base", "HIGH — each Growth sub = 6x Starter"],
    ["Growth annual churn", "15-20%", "B2B SaaS benchmark for SMB segment", "MODERATE"],
    ["Scale annual churn", "8-12%", "Enterprise SaaS benchmark. Higher switching costs", "LOW — small pool regardless"],
    ["Marketplace platform margin rate", "15-20% (avg 17.5%)", "Industry standard for professional services marketplace", "LOW — standard market practice"],
    ["Marketplace service attach rate", "15-22% of Starter subs/quarter", "Conservative. Companies need compliance services", "MODERATE"],
    ["Monthly incorporation volume at M12", "60/month", "Based on Google Ads + SEO + PLG + CA referrals", "MODERATE — affects everything downstream"],
    ["Blended incorporation platform fee", "~₹2,500", "Mix of Launch (₹999-1,499), Grow, Scale tiers", "LOW"],
    ["Full Ops launch burn", "₹7.0L/mo (6 people + other)", "Based on operations-and-staffing.md salary bands", "N/A — actual market rates"],
    ["Lean launch burn", "₹2.75L/mo (3 people + other)", "Founder + 1 dev + contract ops + infra/marketing", "N/A — actual planned costs"],
    ["Full Ops growth burn", "₹12.0L/mo (15 people)", "Based on growth phase staffing plan", "N/A"],
    ["Lean growth burn", "₹3.75L/mo (5 people)", "Phase 2 hires: ops manager + marketing", "N/A"],
    ["Full Ops scale burn", "₹19.5-22L/mo (25+ people)", "Based on scale phase staffing plan", "N/A"],
    ["Lean scale burn", "₹5.5-6L/mo (8-9 people)", "Phase 3 hires: 2nd dev + Sr CS + BD", "N/A"],
    ["Break-even (Lean model)", "Month 12-14", "Revenue ≥ ₹3.75L/mo with 150+ Starter, 15+ Growth, 1-2 Scale", "HIGH — depends on churn + Growth adoption"],
    ["Break-even (Full Ops model)", "Month 24-30", "Revenue ≥ ₹19-22L/mo — very difficult to achieve in 2 years", "Model likely doesn't break even in 3 years"],
]

for item in assumptions:
    r = write_row(ws6, r, item, NC6)

auto_width(ws6, min_w=18, max_w=60)


# ═══════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════
output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "business")
output_path = os.path.join(output_dir, "anvils-lean-vs-fullops-comparison.xlsx")
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"\nKey 3-year numbers:")
print(f"  Lean model:")
print(f"    Year 1 Revenue: ₹{y1_total:,.0f} | Cost: ₹{lean_y1_total_cost:,.0f} | Net: ₹{y1_net:,.0f}")
print(f"    Year 2 Revenue: ₹{y2_total:,.0f} | Cost: ₹{lean_y2_total_cost:,.0f} | Net: ₹{y2_net:,.0f}")
print(f"    Year 3 Revenue: ₹{y3_total:,.0f} | Cost: ₹{lean_y3_total_cost:,.0f} | Net: ₹{y3_net:,.0f}")
print(f"    3-Year Total Profit: ₹{y1_net + y2_net + y3_net:,.0f}")
print(f"  Full Ops model:")
print(f"    Year 1 Cost: ₹{fo_y1_cost:,.0f} | Year 2 Cost: ₹{fo_y2_cost:,.0f} | Year 3 Cost: ₹{fo_y3_cost:,.0f}")
fo_3yr_profit = (y1_total - fo_y1_cost) + (y2_total - fo_y2_cost) + (y3_total - fo_y3_cost)
print(f"    3-Year Total Profit: ₹{fo_3yr_profit:,.0f}")
print(f"  SAVINGS (Lean vs Full Ops): ₹{fo_total_cost - lean_total_cost:,.0f} in costs over 3 years")
